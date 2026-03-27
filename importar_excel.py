"""
Importa registros de gastos desde un archivo Excel al sistema EPG.
Uso: python3.10 importar_excel.py nombre_del_archivo.xlsx

El Excel debe tener estas columnas (en cualquier orden):
  MES, FACULTAD/UNIDAD (o FACULTAD), MENCION, EXPEDIENTE, OFICIO,
  DESCRIPCION, MONTO, ESTADO, OBSERVACION, CONDICION
"""
import sys
import os

# Mapeo flexible de nombres de columna
COLUMNAS = {
    'mes':         ['MES'],
    'facultad':    ['FACULTAD/UNIDAD', 'FACULTAD', 'FACULTAD/UNIDAD '],
    'mencion':     ['MENCION', 'MENCIÓN'],
    'expediente':  ['EXPEDIENTE', 'N° EXPEDIENTE', 'NRO EXPEDIENTE'],
    'oficio':      ['OFICIO', 'N° OFICIO', 'NRO OFICIO'],
    'descripcion': ['DESCRIPCION', 'DESCRIPCIÓN', 'TIPO'],
    'monto':       ['MONTO', 'IMPORTE', 'MONTO S/'],
    'estado':      ['ESTADO'],
    'observacion': ['OBSERVACION', 'OBSERVACIÓN', 'OBS'],
    'condicion':   ['CONDICION', 'CONDICIÓN'],
    'anio':        ['AÑO', 'ANIO', 'ANO'],
}


def encontrar_col(headers, opciones):
    """Busca el índice de columna probando varios nombres posibles."""
    for op in opciones:
        for i, h in enumerate(headers):
            if str(h).strip().upper() == op.upper():
                return i
    return None


def importar(archivo_excel):
    from openpyxl import load_workbook
    from app import create_app
    from extensions import db
    from models.registro_gasto import RegistroGasto
    from models.programa import Programa
    from datetime import datetime

    app = create_app()

    wb = load_workbook(archivo_excel, data_only=True)
    ws = wb.active

    # Leer encabezados (primera fila no vacía)
    headers = []
    header_row = 1
    for row in ws.iter_rows(min_row=1, max_row=5):
        vals = [str(c.value).strip() if c.value else '' for c in row]
        if any(v for v in vals):
            headers = vals
            header_row = row[0].row
            break

    print(f"Encabezados encontrados en fila {header_row}: {headers}")

    # Mapear columnas
    idx = {}
    for campo, opciones in COLUMNAS.items():
        idx[campo] = encontrar_col(headers, opciones)

    print("\nMapeo de columnas:")
    for k, v in idx.items():
        print(f"  {k:15} -> columna {v} ({headers[v] if v is not None else 'NO ENCONTRADA'})")

    campos_requeridos = ['mes', 'facultad', 'mencion', 'descripcion', 'monto']
    faltantes = [c for c in campos_requeridos if idx.get(c) is None]
    if faltantes:
        print(f"\nERROR: No se encontraron columnas requeridas: {faltantes}")
        print("Revisa que el Excel tenga esas columnas.")
        return

    with app.app_context():
        insertados = 0
        omitidos = 0
        errores = 0

        for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
            # Saltar filas vacías
            if not any(row):
                continue

            def get(campo, default=''):
                i = idx.get(campo)
                if i is None:
                    return default
                val = row[i]
                return str(val).strip() if val is not None else default

            mes         = get('mes').upper()
            facultad    = get('facultad')
            mencion     = get('mencion')
            expediente  = get('expediente') or 'S/N'
            oficio      = get('oficio')
            descripcion = get('descripcion').upper()
            observacion = get('observacion')
            condicion   = get('condicion', 'REALIZADO').upper() or 'REALIZADO'
            estado      = get('estado', 'APROBADO').upper() or 'APROBADO'

            # Anio
            anio_raw = get('anio', '2026')
            try:
                anio = int(float(anio_raw)) if anio_raw else 2026
            except:
                anio = 2026

            # Monto
            try:
                monto_raw = row[idx['monto']]
                monto = float(str(monto_raw).replace('S/', '').replace(',', '').strip())
            except:
                print(f"  [!] Fila con monto invalido: {row}")
                errores += 1
                continue

            if not mes or not mencion or monto <= 0:
                omitidos += 1
                continue

            # Verificar que la mencion exista en programas
            prog = Programa.query.filter_by(mencion=mencion).first()
            if not prog:
                # Buscar por coincidencia parcial
                prog = Programa.query.filter(
                    Programa.mencion.ilike(f'%{mencion[:30]}%')
                ).first()
                if prog:
                    mencion = prog.mencion  # usar nombre exacto
                    facultad = prog.facultad
                else:
                    print(f"  [!] Mencion no encontrada: {mencion[:60]}")
                    errores += 1
                    continue

            reg = RegistroGasto(
                mes=mes,
                anio=anio,
                facultad=facultad or prog.facultad,
                mencion=mencion,
                expediente=expediente,
                oficio=oficio,
                descripcion=descripcion,
                monto=monto,
                estado=estado,
                observacion=observacion,
                condicion=condicion,
                registrado_por='importacion_excel',
                fecha_registro=datetime.utcnow()
            )
            db.session.add(reg)
            insertados += 1

            if insertados % 50 == 0:
                db.session.commit()
                print(f"  ... {insertados} registros procesados")

        db.session.commit()
        print(f"\nImportacion completada:")
        print(f"  Insertados: {insertados}")
        print(f"  Omitidos:   {omitidos}")
        print(f"  Errores:    {errores}")


if __name__ == '__main__':
    if len(sys.argv) < 2:
        print("Uso: python3.10 importar_excel.py archivo.xlsx")
        print("Ejemplo: python3.10 importar_excel.py planillas_febrero.xlsx")
        sys.exit(1)

    archivo = sys.argv[1]
    if not os.path.exists(archivo):
        print(f"ERROR: No se encuentra el archivo '{archivo}'")
        sys.exit(1)

    print(f"Importando desde: {archivo}")
    importar(archivo)
