from flask import Blueprint, render_template, request, jsonify, flash, redirect, url_for, send_file
from flask_login import login_required, current_user
from functools import wraps
from extensions import db
from models.programa import Programa
from models.registro_gasto import RegistroGasto
from models.historial_edicion import HistorialEdicion
from datetime import datetime
import io

operador = Blueprint('operador', __name__)

MESES = ['ENERO', 'FEBRERO', 'MARZO', 'ABRIL', 'MAYO', 'JUNIO',
         'JULIO', 'AGOSTO', 'SEPTIEMBRE', 'OCTUBRE', 'NOVIEMBRE', 'DICIEMBRE']

DESCRIPCIONES = ['HONORARIOS', 'SUBVENCION', 'PLANILLA DIRECTORIO', 'PLANILLA DE SUSTENTACION',
                 'PLANILLA DOCENTE INVITADO', 'PLANILLA DOCENTE NOMBRADO',
                 'POR CONTRATAR DOCENTE', 'OTROS']


def rol_requerido(*roles):
    def decorator(f):
        @wraps(f)
        def decorated(*args, **kwargs):
            if current_user.rol not in roles:
                flash('No tienes permiso para acceder a esta seccion.', 'danger')
                return redirect(url_for('auth.login'))
            return f(*args, **kwargs)
        return decorated
    return decorator


@operador.route('/operador/registro', methods=['GET', 'POST'])
@login_required
@rol_requerido('operador', 'admin')
def registro():
    facultades = db.session.query(Programa.facultad).distinct().order_by(Programa.facultad).all()
    facultades = [f[0] for f in facultades]

    if request.method == 'POST':
        mes = request.form.get('mes')
        anio = int(request.form.get('anio', datetime.now().year))
        facultad = request.form.get('facultad')
        mencion = request.form.get('mencion')
        expediente = request.form.get('expediente', '').strip() or 'S/N'
        oficio = request.form.get('oficio', '').strip()
        descripcion = request.form.get('descripcion')
        observacion = request.form.get('observacion', '').strip()

        es_por_contratar = descripcion == 'POR CONTRATAR DOCENTE'
        monto = 0.0 if es_por_contratar else float(request.form.get('monto', 0) or 0)

        # Validar duplicado solo cuando expediente y oficio esten ambos llenos
        if expediente and expediente != 'S/N' and oficio:
            dup = RegistroGasto.query.filter_by(expediente=expediente, oficio=oficio).first()
            if dup:
                flash(
                    f'No se puede registrar: ya existe el registro #{dup.id} con el mismo expediente y oficio. '
                    f'Buscalo en la primera columna (#) de la lista. '
                    f'Detalle: {dup.mes} {dup.anio} - {dup.facultad} - {dup.mencion} - '
                    f'{dup.descripcion} - S/ {dup.monto:,.2f}.',
                    'danger'
                )
                return redirect(url_for('operador.registro'))

        if es_por_contratar:
            estado = 'APROBADO'
            condicion = 'POR CONTRATAR'
        else:
            prog = Programa.query.filter_by(mencion=mencion).first()
            if not prog:
                flash('Mencion no encontrada.', 'danger')
                return redirect(url_for('operador.registro'))
            if prog.saldo_actual >= monto:
                estado = 'APROBADO'
            else:
                flash('Sin saldo disponible. El registro no puede guardarse.', 'danger')
                return redirect(url_for('operador.registro'))
            condicion = 'EN PROCESO'

        reg = RegistroGasto(
            mes=mes, anio=anio, facultad=facultad, mencion=mencion,
            expediente=expediente, oficio=oficio, descripcion=descripcion,
            monto=monto, estado=estado, observacion=observacion,
            condicion=condicion, registrado_por=current_user.username
        )
        db.session.add(reg)
        db.session.commit()
        flash('Gasto registrado correctamente.', 'success')
        return redirect(url_for('operador.lista'))

    return render_template('operador/registro.html',
                           facultades=facultades,
                           meses=MESES,
                           descripciones=DESCRIPCIONES,
                           anio_actual=datetime.now().year)


@operador.route('/operador/saldo_tiempo_real')
@login_required
@rol_requerido('operador', 'admin')
def saldo_tiempo_real():
    mencion = request.args.get('mencion', '')
    monto = float(request.args.get('monto', 0))
    prog = Programa.query.filter_by(mencion=mencion).first()
    if not prog:
        return jsonify({'ok': False, 'mensaje': 'Mencion no encontrada'})

    saldo_disp = prog.saldo_actual
    aprobado = saldo_disp >= monto
    return jsonify({
        'ok': True,
        'saldo_disponible': saldo_disp,
        'aprobado': aprobado,
        'ingresos': prog.ingresos,
        'gastos': prog.gastos_total,
        'retencion_ocep': prog.retencion_ocep,
        'retencion_epg': prog.retencion_epg
    })


@operador.route('/operador/menciones_por_facultad')
@login_required
def menciones_por_facultad():
    facultad = request.args.get('facultad', '')
    programas = Programa.query.filter_by(facultad=facultad).order_by(Programa.mencion).all()
    return jsonify([{'mencion': p.mencion} for p in programas])


@operador.route('/operador/verificar_duplicado')
@login_required
@rol_requerido('operador', 'admin')
def verificar_duplicado():
    expediente = request.args.get('expediente', '').strip()
    oficio = request.args.get('oficio', '').strip()
    if not expediente or expediente == 'S/N' or not oficio:
        return jsonify({'duplicado': False})
    dup = RegistroGasto.query.filter_by(expediente=expediente, oficio=oficio).first()
    if not dup:
        return jsonify({'duplicado': False})
    return jsonify({
        'duplicado': True,
        'id': dup.id,
        'mes': dup.mes,
        'anio': dup.anio,
        'facultad': dup.facultad,
        'mencion': dup.mencion,
        'descripcion': dup.descripcion,
        'monto': dup.monto,
    })


@operador.route('/operador/lista')
@login_required
@rol_requerido('operador', 'admin')
def lista():
    page = request.args.get('page', 1, type=int)
    per_page = 20

    query = RegistroGasto.query

    # Filtros
    f_mes = request.args.get('mes', '')
    f_facultad = request.args.get('facultad', '')
    f_mencion = request.args.get('mencion', '').strip()
    f_expediente = request.args.get('expediente', '').strip()
    f_oficio = request.args.get('oficio', '').strip()
    f_descripcion = request.args.get('descripcion', '')
    f_estado = request.args.get('estado', '')
    f_condicion = request.args.get('condicion', '')
    f_anio = request.args.get('anio', '')

    if f_mes:
        query = query.filter(RegistroGasto.mes == f_mes)
    if f_facultad:
        query = query.filter(RegistroGasto.facultad == f_facultad)
    if f_mencion:
        query = query.filter(RegistroGasto.mencion.contains(f_mencion))
    if f_expediente:
        query = query.filter(RegistroGasto.expediente.contains(f_expediente))
    if f_oficio:
        query = query.filter(RegistroGasto.oficio.contains(f_oficio))
    if f_descripcion:
        query = query.filter(RegistroGasto.descripcion == f_descripcion)
    if f_estado:
        query = query.filter(RegistroGasto.estado == f_estado)
    if f_condicion:
        query = query.filter(RegistroGasto.condicion == f_condicion)
    if f_anio:
        query = query.filter(RegistroGasto.anio == int(f_anio))

    total_monto = query.with_entities(db.func.sum(RegistroGasto.monto)).scalar() or 0.0
    total_registros = query.count()

    registros = query.order_by(RegistroGasto.fecha_registro.desc()).paginate(
        page=page, per_page=per_page, error_out=False)

    facultades = db.session.query(Programa.facultad).distinct().order_by(Programa.facultad).all()

    # Opciones por facultad para los datalists de busqueda
    menciones_data = {}
    for p in Programa.query.order_by(Programa.facultad, Programa.mencion).all():
        menciones_data.setdefault(p.facultad, []).append(p.mencion)

    expedientes_data, oficios_data = {}, {}
    rows = db.session.query(
        RegistroGasto.facultad, RegistroGasto.expediente, RegistroGasto.oficio
    ).distinct().all()
    for fac, exp, ofi in rows:
        if exp:
            expedientes_data.setdefault(fac, set()).add(exp)
        if ofi:
            oficios_data.setdefault(fac, set()).add(ofi)
    expedientes_data = {k: sorted(v) for k, v in expedientes_data.items()}
    oficios_data = {k: sorted(v) for k, v in oficios_data.items()}

    return render_template('operador/lista.html',
                           registros=registros,
                           total_monto=total_monto,
                           total_registros=total_registros,
                           facultades=[f[0] for f in facultades],
                           meses=MESES,
                           descripciones=DESCRIPCIONES,
                           menciones_data=menciones_data,
                           expedientes_data=expedientes_data,
                           oficios_data=oficios_data,
                           filtros={
                               'mes': f_mes, 'facultad': f_facultad,
                               'mencion': f_mencion, 'expediente': f_expediente,
                               'oficio': f_oficio, 'descripcion': f_descripcion,
                               'estado': f_estado, 'condicion': f_condicion, 'anio': f_anio
                           })


CAMPOS_EDITABLES = ['mes', 'anio', 'facultad', 'mencion', 'expediente', 'oficio',
                    'descripcion', 'estado', 'condicion', 'observacion']


def _aplicar_edicion(reg, form, user, rol):
    """Actualiza campos de reg desde form (excluye monto) y registra historial."""
    cambios = []
    for campo in CAMPOS_EDITABLES:
        if campo not in form:
            continue
        nuevo = form.get(campo, '').strip()
        if campo == 'anio':
            nuevo_val = int(nuevo) if nuevo else reg.anio
        elif campo in ('mes', 'descripcion', 'estado', 'condicion'):
            nuevo_val = nuevo.upper() if nuevo else getattr(reg, campo)
        elif campo == 'expediente':
            nuevo_val = nuevo or 'S/N'
        else:
            nuevo_val = nuevo
        anterior = getattr(reg, campo)
        if (anterior or '') != (nuevo_val or ''):
            cambios.append((campo, anterior, nuevo_val))
            setattr(reg, campo, nuevo_val)
    for campo, ant, nue in cambios:
        db.session.add(HistorialEdicion(
            registro_id=reg.id, editado_por=user, rol=rol,
            campo=campo, valor_anterior=str(ant) if ant is not None else '',
            valor_nuevo=str(nue) if nue is not None else '',
            expediente=reg.expediente
        ))
    return len(cambios)


def _aplicar_transicion_por_contratar(reg, form, user, rol):
    """Si el registro es POR CONTRATAR DOCENTE y se cambia a otra descripcion,
    valida monto y saldo. Devuelve (error_msg, forzar_en_proceso)."""
    if reg.descripcion != 'POR CONTRATAR DOCENTE':
        return None, False
    nueva_desc = (form.get('descripcion') or '').strip().upper()
    if not nueva_desc or nueva_desc == 'POR CONTRATAR DOCENTE':
        return None, False
    form_monto = (form.get('monto') or '').strip()
    try:
        nuevo_monto = float(form_monto)
    except ValueError:
        nuevo_monto = 0.0
    if nuevo_monto <= 0:
        return 'Al cambiar la descripcion debes ingresar el monto real (mayor a 0).', False
    prog = Programa.query.filter_by(mencion=reg.mencion).first()
    if not prog or prog.saldo_actual < nuevo_monto:
        return 'Sin saldo disponible para activar este registro con el monto indicado.', False
    if reg.monto != nuevo_monto:
        db.session.add(HistorialEdicion(
            registro_id=reg.id, editado_por=user, rol=rol,
            campo='monto', valor_anterior=str(reg.monto), valor_nuevo=str(nuevo_monto),
            expediente=reg.expediente
        ))
        reg.monto = nuevo_monto
    return None, True


@operador.route('/operador/editar/<int:reg_id>', methods=['POST'])
@login_required
@rol_requerido('operador', 'admin')
def editar_registro(reg_id):
    reg = RegistroGasto.query.get_or_404(reg_id)
    if reg.condicion == 'REALIZADO':
        flash('No se puede editar un registro en condicion REALIZADO.', 'danger')
        return redirect(request.referrer or url_for('operador.lista'))
    error, forzar_en_proceso = _aplicar_transicion_por_contratar(
        reg, request.form, current_user.username, current_user.rol)
    if error:
        flash(error, 'danger')
        return redirect(request.referrer or url_for('operador.lista'))
    n = _aplicar_edicion(reg, request.form, current_user.username, current_user.rol)
    if forzar_en_proceso and reg.condicion in ('POR CONTRATAR', ''):
        reg.condicion = 'EN PROCESO'
    db.session.commit()
    flash(f'Registro actualizado ({n} cambio(s)).' if n else 'Sin cambios.', 'success' if n else 'info')
    return redirect(request.referrer or url_for('operador.lista'))


@operador.route('/operador/eliminar/<int:reg_id>', methods=['POST'])
@login_required
@rol_requerido('operador', 'admin')
def eliminar_registro(reg_id):
    reg = RegistroGasto.query.get_or_404(reg_id)
    info = f'Exp. {reg.expediente} - {reg.descripcion} - S/ {reg.monto:,.2f}'
    db.session.delete(reg)
    db.session.commit()
    flash(f'Registro eliminado: {info}', 'success')
    return redirect(request.referrer or url_for('operador.lista'))


@operador.route('/operador/exportar_registros')
@login_required
@rol_requerido('operador', 'admin')
def exportar_registros():
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Registros de Gastos'

    header_fill = PatternFill(start_color='960000', end_color='960000', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True)
    thin = Side(style='thin')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    headers = ['#', 'MES', 'AÑO', 'FACULTAD', 'MENCION', 'EXPEDIENTE', 'OFICIO',
               'DESCRIPCION', 'MONTO', 'ESTADO', 'CONDICION', 'OBSERVACION',
               'REGISTRADO POR', 'FECHA REGISTRO']

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
        cell.border = border

    registros = RegistroGasto.query.order_by(RegistroGasto.fecha_registro.desc()).all()
    for i, r in enumerate(registros, 1):
        row = i + 1
        datos = [i, r.mes, r.anio, r.facultad, r.mencion, r.expediente, r.oficio,
                 r.descripcion, r.monto, r.estado, r.condicion, r.observacion,
                 r.registrado_por,
                 r.fecha_registro.strftime('%d/%m/%Y %H:%M') if r.fecha_registro else '']
        for col, val in enumerate(datos, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.border = border
            if col == 9:
                cell.number_format = '#,##0.00'

    ws.column_dimensions['D'].width = 30
    ws.column_dimensions['E'].width = 50
    ws.column_dimensions['H'].width = 30

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(output, download_name='registros_gastos.xlsx',
                     as_attachment=True,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
