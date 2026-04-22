from flask import Blueprint, render_template, request, jsonify, send_file, flash, redirect, url_for
from flask_login import login_required, current_user
from functools import wraps
from extensions import db
from models.programa import Programa, HistorialIngreso
import io

tesorera = Blueprint('tesorera', __name__)


def rol_requerido(*roles):
    def decorator(f):
        @wraps(f)
        def decorated(*args, **kwargs):
            if current_user.rol not in roles:
                flash('No tienes permiso para acceder a esta seccion.', 'danger')
                return redirect(url_for('auth.login'))
            return f(*args, **kwargs)
        return decorated
    return decorator


@tesorera.route('/tesorera/presupuesto')
@login_required
@rol_requerido('tesorera', 'admin')
def presupuesto():
    programas = Programa.query.order_by(Programa.facultad, Programa.tipo_programa).all()
    return render_template('tesorera/presupuesto.html', programas=programas)


@tesorera.route('/tesorera/actualizar_ingreso/<int:prog_id>', methods=['POST'])
@login_required
@rol_requerido('tesorera', 'admin')
def actualizar_ingreso(prog_id):
    prog = Programa.query.get_or_404(prog_id)
    data = request.get_json()
    nuevo_monto = float(data.get('ingresos', 0))

    historial = HistorialIngreso(
        programa_id=prog.id,
        monto_anterior=prog.ingresos,
        monto_nuevo=nuevo_monto,
        modificado_por=current_user.username
    )
    db.session.add(historial)
    prog.ingresos = nuevo_monto
    db.session.commit()

    return jsonify({
        'ok': True,
        'gastos': prog.gastos_total,
        'retencion_ocep': prog.retencion_ocep,
        'retencion_epg': prog.retencion_epg,
        'saldo': prog.saldo_actual,
        'situacion': prog.situacion
    })


@tesorera.route('/tesorera/exportar')
@login_required
@rol_requerido('tesorera', 'admin')
def exportar():
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Presupuesto EPG'

    header_fill = PatternFill(start_color='960000', end_color='960000', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True)
    thin = Side(style='thin')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    headers = ['#', 'TIPO', 'FACULTAD', 'MENCION', 'INGRESOS', 'GASTOS',
               'RET. OCEP (10%)', 'RET. EPG (5%)', 'SALDO', 'SITUACION']
    col_widths = [5, 12, 35, 55, 15, 15, 16, 16, 15, 15]

    for col, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = border
        ws.column_dimensions[cell.column_letter].width = w

    ws.row_dimensions[1].height = 30

    fills = {
        'BIEN': PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid'),
        'CRITICO': PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid'),
        'EN EL LIMITE': PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid'),
        'SOBREPASADO': PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid'),
    }

    programas = Programa.query.order_by(Programa.facultad, Programa.tipo_programa).all()
    for i, p in enumerate(programas, 1):
        row = i + 1
        datos = [i, p.tipo_programa, p.facultad, p.mencion,
                 p.ingresos, p.gastos_total, p.retencion_ocep, p.retencion_epg,
                 p.saldo_actual, p.situacion]
        for col, val in enumerate(datos, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.border = border
            if col in [5, 6, 7, 8, 9]:
                cell.number_format = '#,##0.00'
                cell.alignment = Alignment(horizontal='right')
            if col == 10:
                sit = p.situacion
                if sit in fills:
                    cell.fill = fills[sit]
                cell.alignment = Alignment(horizontal='center')

    # Totales
    last = len(programas) + 2
    ws.cell(row=last, column=4, value='TOTALES').font = Font(bold=True)
    for col in [5, 6, 7, 8, 9]:
        cell = ws.cell(row=last, column=col)
        letter = cell.column_letter
        cell.value = f'=SUM({letter}2:{letter}{last-1})'
        cell.number_format = '#,##0.00'
        cell.font = Font(bold=True)
        cell.border = border

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(output, download_name='presupuesto_epg.xlsx',
                     as_attachment=True,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
